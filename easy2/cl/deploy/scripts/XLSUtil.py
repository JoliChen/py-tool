# -*- coding: utf-8 -*-
# @Time    : 2021/5/11 11:15 AM
# @Author  : Joli
# @Email   : 99755349@qq.com

# dependences
# pip install openpyxl

import json
import os
import sys
import openpyxl

EXTENSION = '.xlsx'
END = 'END'

TYPE_BOOL = 1
TYPE_INT = 2
TYPE_LONG = 3
TYPE_FLOAT = 4
TYPE_STRING = 5
ALL_TYPE_DICT = {'bool': TYPE_BOOL, 'int': TYPE_INT, 'long': TYPE_LONG, 'float': TYPE_FLOAT, 'string': TYPE_STRING}

def to_int(v):
    try:
        return int(v)
    except (ValueError, TypeError):
        pass
    return 0

def to_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        pass
    return 0.0

def to_string(v):
    return v.strip() if isinstance(v, str) else ''

def guess_value(v, t):
    if t == TYPE_STRING:
        return to_string(v)
    if t == TYPE_INT or t == TYPE_LONG:
        return to_int(v)
    if t == TYPE_FLOAT:
        return to_float(v)
    return v

def guess_type(v):
    if isinstance(v, str):
        return 'string'
    elif isinstance(v, int):
        if v > 2147483647 or v < -2147483648:
            return 'long'
        else:
            return 'int'
    elif isinstance(v, float):
        return 'float'
    elif isinstance(v, bool):
        return 'bool'
    elif isinstance(v, list):
        # return 'list'
        return 'string'
    elif isinstance(v, dict):
        # return 'dict'
        return 'string'
    return ''

def compare_string(s1, s2):
    n1, n2 = len(s1), len(s2)
    for i in range(min(n1, n2)):
        c1, c2 = ord(s1[i]), ord(s2[i])
        if c1 > c2:
            return 1
        if c1 < c2:
            return -1
    if n1 > n2:
        return 1
    if n1 < n2:
        return -1
    return 0

def compare_number(i1, i2):
    if i1 > i2:
        return 1
    if i1 < i2:
        return -1
    return 0

def compare_value(v1, v2, t):
    if t == TYPE_STRING:
        return compare_string(v1, v2)
    return compare_number(v1, v2)

def find_cell_in_row(ws, row, field):
    for c in range(ws.max_column):
        cell = ws[row][c]
        if cell.value == field:
            return cell

def find_cell_by_field(ws, field):
    for r in range(1, ws.max_row):  # 从1开始
        for c in range(ws.max_column):  # 从0开始
            cell = ws[r][c]
            if cell.value == field:
                return cell

def find_client_cell(ws):
    return find_cell_by_field(ws, 'CLIENT')

def find_server_cell(ws):
    return find_cell_by_field(ws, 'SERVER')

def find_primary_cell(ws):
    cell = find_client_cell(ws)
    if cell:
        idcell = ws[cell.row][cell.col_idx]  # cell.row 和 cell.col_idx 是从1开始的
        return idcell

def check_insert_row(ws, idcell, rowdata):
    newid = rowdata.get(idcell.value)
    r, c = idcell.row, idcell.col_idx-1
    endc = c - 1
    nrow = ws.max_row
    vals = {}
    idtype = ALL_TYPE_DICT[ws[r + 1][c].value]
    for i in range(r + 3, nrow):
        v = guess_value(ws[i][c].value, idtype)
        vals[i] = v
        if newid == v:
            return i
        if ws[i][endc].value == END:
            break
    nr, hasend = nrow, False
    for i in range(r + 3, nrow):
        if compare_value(newid, vals[i], idtype) < 0:
            nr = i
            break
        if ws[i][endc].value == END:
            ws[i][endc].value = ''
            hasend = True
            nr = i + 1
            break
    ws.insert_rows(nr)  # 插入一行
    if hasend:
        ws[nr][endc].value = END
    return nr

def check_insert_col(ws, idcell, field, value):
    r = idcell.row
    cell = find_cell_in_row(ws, r, field)
    if cell:
        return cell.col_idx - 1
    c, tr = ws.max_column, r - 1
    for i in range(idcell.col_idx, ws.max_column):
        if not ws[tr][i].value:
            c = i + 1
            break
    ws.insert_cols(c)  # 插入一列
    c = c - 1
    ws[r][c].value = field
    ws[r + 1][c].value = guess_type(value)
    ws[r + 2][c].value = field
    return c

def write_sheet(ws, wsdata):
    idcell = find_primary_cell(ws)
    field_cols = {}
    for rowdata in wsdata:
        r = check_insert_row(ws, idcell, rowdata)
        for field, value in rowdata.items():
            if field in field_cols:
                c = field_cols[field]
            else:
                c = check_insert_col(ws, idcell, field, value)
                field_cols[field] = c
            cell = ws[r][c]
            cell.value = value

def read_sheet_all(ws):
    wsdata = []
    for cols in ws.rows:
        rowdata = []
        for cell in cols:
            rowdata.append(cell.value)
        wsdata.append(rowdata)
    return wsdata

def write(jsondir, xlsxdir, wballows: list=None, wsallows: list=None):
    for fn in os.listdir(jsondir):
        exts = os.path.splitext(fn)
        if '.json' != exts[1]:
            continue
        wbname = exts[0]
        if wballows and wbname not in wballows:
            continue
        wbfile = os.path.join(xlsxdir, wbname + EXTENSION)
        wb = openpyxl.load_workbook(wbfile)
        if not wb:
            continue
        with open(os.path.join(jsondir, fn), encoding='utf-8') as fp:
            wbdata = json.load(fp)
        for wsname in wbdata:
            if wsallows and wsname not in wsallows:
                continue
            write_sheet(wb[wsname], wbdata[wsname])
        wb.save(wbfile)
    # print('write done', xlsxdir)

def read(xlsxdir, jsondir, wballows: list=None, wsallows: list=None):
    data = {}
    for fn in os.listdir(xlsxdir):
        if fn.startswith('~'):
            continue
        exts = os.path.splitext(fn)
        if EXTENSION != exts[1]:
            continue
        wbname = exts[0]
        if wballows and wbname not in wballows:
            continue
        wbdata = data.setdefault(wbname, {})
        wb = openpyxl.load_workbook(os.path.join(xlsxdir, fn), read_only=True)
        for wsname in wb.sheetnames:
            if wsallows and wsname not in wsallows:
                continue
            wbdata[wsname] = read_sheet_all(wb[wsname])
    if jsondir:
        if not os.path.isdir(jsondir):
            os.makedirs(jsondir)
        for exname, table in data.items():
            fn = os.path.join(jsondir, exname + '.json')
            with open(fn, 'w', encoding='utf-8') as fp:
                json.dump(table, fp, ensure_ascii=False, indent='\t')
    else:
        s = json.dumps(data, ensure_ascii=False, indent='\t')
        sys.stdout.write(s)