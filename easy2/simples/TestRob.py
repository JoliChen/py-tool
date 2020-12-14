# -*- coding: utf-8 -*-
# @Time    : 2019/7/6 4:35 PM
# @Author  : Joli
# @Email   : 99755349@qq.com
import json
import os
import shutil
import zlib
import requests
from jonlin.utils import FS, Log, Bit
from jonlin.cl import Shell

log = Log.Logger(__file__)

def test_load_cdn_files():
    dstdir = '/Users/joli/Desktop/dldlh5/cdn'
    cdnurl = 'https://res.xxh5.z7xz.com/xxh5dev'
    errors = []

    # 根据资源表加载资源
    def load_by_json():
        flist = json.loads(FS.read_text('/Users/joli/Desktop/dldlh5/files.json'))
        httpss = requests.session()
        httpss.keep_alive = False

        def _httpload(host, filename):
            # filename = '%s_%s%s' % (FS.get_file_name(key), obj, ext)
            file_url = '%s/%s' % (host, filename)
            response = httpss.get(file_url)
            if response.status_code == 200:
                filename = file_url[len(cdnurl) + 1:]
                print(filename)
                filepath = os.path.join(dstdir, filename)
                FS.make_parent(filepath)
                with open(filepath, "wb") as fp:
                    fp.write(response.content)
                return True

        def fullpath(parent, info):
            for key in info:
                obj = info[key]
                if isinstance(obj, dict):
                    fullpath('%s/%s' % (parent, key), obj)
                else:
                    ext = FS.extensions(key)
                    isetc = False
                    if ext == '.etc':
                        isetc = True
                        ext = '.png'
                    filename = '%s_%s%s' % (FS.filename(key), obj, ext)
                    if not _httpload(parent, filename):
                        if isetc:
                            filename = '%s_%s%s' % (FS.filename(key), obj, '.jpg')
                            if not _httpload(parent, filename):
                                filename = '%s_%s%s' % (FS.filename(key), obj, '.etc')
                                if not _httpload(parent, filename):
                                    failurl = parent + '/' + filename
                                    log.e('fail', failurl)
                                    errors.append(failurl)
                        else:
                            failurl = parent + '/' + filename
                            log.e('fail', failurl)
                            errors.append(failurl)
        flist.pop('skins')
        fullpath(cdnurl, flist)

    # 根据文件目录结构加载资源
    def load_by_disktree():
        svdir = '/Users/joli/Desktop/apps/h5/map'
        iddir = os.path.join(dstdir, 'map')
        flist = sorted(os.listdir(iddir))
        guess = 1000
        ss = requests.session()
        ss.keep_alive = False
        num = 0
        # for n in range(len(flist)-1, -1, -1):
        for n in range(95, -1, -1):
            fid = flist[n]
            col = guess
            end = False
            for i in range(guess):
                for j in range(col):
                    filename = 'pic%d_%d.swf' % (i, j)
                    file_url = '%s/map/%s/swf/%s' % (cdnurl, fid, filename)
                    log.d(n, "load", file_url)
                    response = ss.get(file_url)
                    if response.status_code == 404:
                        if i == 0:
                            col = j
                        else:
                            end = True
                        break
                    elif response.status_code == 200:
                        filepath = os.path.join(svdir, fid, filename)
                        FS.make_parent(filepath)
                        with open(filepath, "wb") as fp:
                            fp.write(response.content)
                    else:
                        log.e(num, file_url)
                        errors.append(file_url)
                if end:
                    break
    # load_by_disktree()
    load_by_json()
    print('----------------------------------', len(errors))
    for url in errors:
        print(url)

def test_hack_xxtea():
    import xxtea

    def fix_key(key):
        size = len(key)
        if size < 16:
            return key + bytes(16 - size)
        elif size > 16:
            return key[:16]
        return key

    def dexxtea(buffer, bsign, bkey):
        sign_len = len(bsign)
        if buffer[0: sign_len] != bsign:
            return buffer
        source = buffer[sign_len:]
        return xxtea.decrypt(source, bkey, padding=False)

    def dexxtea_file(src, dst, bsign, bkey):
        with open(src, 'rb') as fp:
            data = dexxtea(fp.read(), bsign, bkey)
        FS.make_parent(dst)
        with open(dst, 'wb') as fp:
            fp.write(data)

    sjldt = {
        'xxtea_sign': b'Cokutau',
        'xxtea_key': fix_key(b'Cokutau_pjzz'),
        'src_dir': '/Users/joli/Desktop/apps/神将乱斗团/assets/res',
        'dst_dir': '/Users/joli/Desktop/apps/1/神将乱斗团/res',
        'testfile': 'battle/unit_text/cirt_sp.png'
    }
    sszg = {
        'xxtea_sign': b'shiyuegame',
        'xxtea_key': fix_key(b'JYDYdpyRecDuHQc8'),
        'src_dir': '/Users/joli/Desktop/apps/闪烁之光/assets/res',
        'dst_dir': '/Users/joli/Desktop/apps/1/闪烁之光/res',
        'testfile': 'resource/face/face.png'
    }
    qqsg = {
        'xxtea_sign': b'@bianfeng',
        'xxtea_key':  fix_key(b'8x120161031*#0b@^4(j&b='),  # fix_key(b"|-_Rrv2b>LP%4]'rn"),
        'src_dir': '/Users/joli/Desktop/test/hack/com.tencent.tmgp.bfqqsg_1.6.5_110520/assets',
        'dst_dir': '/Users/joli/Desktop/test/hack/com.tencent.tmgp.bfqqsg_1.6.5_110520/assets1',
        'testfile': 'Data/Art/UI_new/bg_activity.jpg'
    }
    target = qqsg
    # for filename in FS.walk_files(target['src_dir'], cut_pos=len(target['src_dir'])+1):
    #     srcpath = os.path.join(target['src_dir'], filename)
    #     dstpath = os.path.join(target['dst_dir'], filename)
    #     print('dexxtea', srcpath)
    #     dexxtea_file(srcpath, dstpath, target['xxtea_sign'], target['xxtea_key'])
    dexxtea_file(
        os.path.join(target['src_dir'], target['testfile']),
        os.path.join(target['dst_dir'], target['testfile']),
        target['xxtea_sign'],
        target['xxtea_key']
    )

def test_hack_xor():
    sig = b'laoma'
    src_dir = '/Users/joli/Downloads/nds_19041501/assets'
    dst_dir = '/Users/joli/Downloads/nds_19041501/de_assets'
    for filename in FS.walk_files(src_dir, cut=len(src_dir) + 1):
        # src = '/Users/joli/Downloads/nds_19041501/assets/src/accountLogin/Img_account_bg.png'
        # dst = '/Users/joli/Downloads/nds_19041501/assets/src/accountLogin/Img_account_bg1.png'
        src = os.path.join(src_dir, filename)
        dst = os.path.join(dst_dir, filename)
        with open(src, 'rb') as fp:
            temp = fp.read()
            if temp[:5] == sig:
                temp = temp[5:]
                size = len(temp)
                data = bytearray(size)
                for i in range(size):
                    data[i] = temp[i] ^ 0x6
            else:
                continue
        FS.make_parent(dst)
        with open(dst, 'wb') as fp:
            fp.write(data)

def test_hack_luac():
    unluac = '/Users/joli/source/mac/app/hack/unluac_2015_06_13.jar'
    src = '/Users/joli/Desktop/apps/apk/res2/script/main.lua'
    dst = '/Users/joli/Desktop/apps/apk/res2/script/main1.lua'
    Shell.run('java -Djava.awt.headless=true -jar %s %s>%s' % (unluac, src, dst))

def test_hack_respack():
    def split_files(pkg, buffer):
        with open(pkg + '_filelist.txt', 'r') as fp:
            filelist = []
            for item in fp.read().split('\n'):
                if item:
                    colon = item.find(':')
                    filelist.append((int(item[0:colon]), item[colon+1:]))
            print("filelist:", len(filelist))
        pos = len(buffer)
        for i in range(len(filelist)-1, -1, -1):
            item = filelist[i]
            path = os.path.join(pkg, item[1])
            FS.make_parent(path)
            beg = item[0] + len(item[1]) + 1
            with open(path, 'wb') as fp:
                fp.write(buffer[beg:pos])
            pos = item[0]

    def parse_package():
        # pkg = '/Users/joli/Desktop/apps/三国杀名将传/assets/package1.assets'
        pkg = '/Users/joli/Desktop/apps/三国杀名将传/assets/patch.assets'
        with open(pkg, 'rb') as fp:
            # extract_extlist(pkg[0:-7], fp.read())
            extract_filelist(pkg[0:-7], fp.read())
            # split_files(pkg[0:-7], fp.read())
            # ba = Bit.ByteArray().init_buffer(buff)
            # print(ba.read_u16())
            # print(ba.read_u16())
            # print(ba.read_u16())
            # print(ba.read_int())
    # parse_package()

    def create_lua_table():
        filelist = set()
        with open('/Users/joli/Desktop/apps/三国杀名将传/assets/package1_filelist.txt', 'r') as fp:
            for item in fp.read().split('\n'):
                if item:
                    name = item[item.find(':')+1:]
                    fext = FS.extensions(name)
                    if fext == '.png' or fext == '.jpg':
                        filelist.add('"%s"' % name)
        with open('/Users/joli/Desktop/apps/三国杀名将传/assets/patch_filelist.txt', 'r') as fp:
            for item in fp.read().split('\n'):
                if item:
                    name = item[item.find(':')+1:]
                    fext = FS.extensions(name)
                    if fext == '.png' or fext == '.jpg':
                        filelist.add('"%s"' % name)
        filelist = sorted(filelist)
        print(len(filelist))
        fileroot = '/Users/joli/Documents/AndroidStudio/DeviceExplorer/emulator-5554/data/user/0/com.tencent.tmgp.sanguosha.mjz/files'
        resroot = '/Users/joli/Desktop/apps/apk/res'
        n = 0
        for i in range(len(filelist)):
            filename = filelist[i][1:-1]
            hackname = 'hackimage_%d.%s' % (i+1, filename[-3:])
            hackpath = os.path.join(fileroot, hackname)
            if os.path.isfile(hackpath):
                dst = os.path.join(resroot, filename)
                FS.make_parent(dst)
                shutil.copy2(hackpath, dst)
            else:
                n += 1
                print("miss", n, i, filename, hackname)
        # print(',\n'.join(filelist))
    create_lua_table()

    # b = bytes([0x9d, 0x02, 0x0, 0x0])
    # print(Bit.u32_from(b))
    # b = bytes([0x9f, 0x02, 0x0, 0x0])
    # print(Bit.u32_from(b))
    # b = bytes([0x45, 0x16, 0x0, 0x0])
    # print(Bit.u32_from(b))
    # b = bytes([0xc9, 0x0, 0x0, 0x0])
    # print(Bit.u32_from(b))
    # print('%x' % 184, '%x' % 69)

def test_hack_zlib():
    def deflate(data, compresslevel=9):
        # compress = zlib.compressobj(
        #     compresslevel,  # level: 0-9
        #     zlib.DEFLATED,  # method: must be DEFLATED
        #     -zlib.MAX_WBITS,  # window size in bits:
        #     #   -15..-8: negate, suppress header
        #     #   8..15: normal
        #     #   16..30: subtract 16, gzip header
        #     zlib.DEF_MEM_LEVEL,  # mem level: 1..8/9
        #     0  # strategy:
        #     #   0 = Z_DEFAULT_STRATEGY
        #     #   1 = Z_FILTERED
        #     #   2 = Z_HUFFMAN_ONLY
        #     #   3 = Z_RLE
        #     #   4 = Z_FIXED
        # )
        encoder = zlib.compressobj()
        buffer = encoder.compress(data)
        buffer += encoder.flush()
        return buffer

    def inflate(data):
        # decoder = zlib.decompressobj(
        #     -zlib.MAX_WBITS  # see above
        # )
        decoder = zlib.decompressobj()
        buffer = decoder.decompress(data)
        buffer += decoder.flush()
        return buffer
        # return zlib.decompress(data)

    src = '/Users/joli/Desktop/dldlh5/filetable_2dw0b3_mix_etc.bin'
    dst = '/Users/joli/Desktop/dldlh5/filetable_2dw0b3_mix_etc_dec.bin'
    # inflate_file(src, dst)
    with open(src, 'rb') as fp:
        bf = fp.read()
        bf = inflate(bf)
        # ba = Bit.ByteArray().init_buffer(bf)
        # ba.set_endian(Bit.BIG_ENDIAN)
        # nums = read_int(ba)
        # print(nums)
        # for n in range(nums):
        #     size = read_int(ba)
        #     print(size)
        #     print(ba.read_utf8(size))
    with open(dst, 'wb') as fp:
        fp.write(bf)

def test_hack_sqlite3():
    pass

def test_json2excel():
    import openpyxl
    import demjson

    def _keysort(v):
        v = v.lower()
        if 'id' == v:
            return 0
        if 'id' in v:
            return 1
        if 'type' == v:
            return 2
        if 'type' in v:
            return 3
        if 'name' == v:
            return 4
        if 'name' in v:
            return 5
        return 6

    def _guess_type(v):
        if isinstance(v, int):
            return 'int'
        if isinstance(v, float):
            return 'float'
        if isinstance(v, str):
            return 'string'
        if isinstance(v, dict):
            return 'dict'
        if isinstance(v, list):
            return 'list'
        raise RuntimeError('unknow type ' + v)

    def _build_sheet1(sheet, config, title, flat=False):  # 要求config是二维数组或二维字典
        sheet.title = title
        ckey = []
        skey = []
        kid1 = {}
        kid2 = []
        isdict = isinstance(config, dict)
        for idx in config:
            if isdict:
                item = config[idx]
            else:
                item = idx
            if not flat:
                for k in item:
                    if k not in ckey:
                        ckey.append(k)
                        kid1[k] = _guess_type(item.get(k))
            else:
                ckey.append(idx)
                kid1[idx] = _guess_type(item)
        ckey.sort(key=_keysort)
        if isdict and not flat and ckey[0] != 'id':
            ckey.insert(0, 'id')
            kid1['id'] = 'int'
            is_add_id = True
        else:
            is_add_id = False
        for k in ckey:
            t = kid1[k]
            if t == 'dict' or t == 'list':
                kid2.append('string')
            else:
                kid2.append(t)
        print(f'{len(ckey)} {ckey}')
        print(f'{len(kid2)} {kid2}')
        # header
        sheet.append([title])
        row = ['CLIENT']
        row.extend(ckey)
        sheet.append(row)
        row = ['']
        row.extend(kid2)
        sheet.append(row)
        row = ['SERVER']
        row.extend(skey)
        sheet.append(row)
        if not flat:
            for idx in config:
                row = ['']
                item = config[idx] if isdict else idx
                if is_add_id:
                    try:
                        row.append(int(idx))
                    except:
                        row.append(idx)
                for i in range(1 if is_add_id else 0, len(ckey)):
                    k = ckey[i]
                    t = kid1[k]
                    v = item.get(k)
                    if t == 'dict' or t == 'list' or isinstance(v, dict) or isinstance(v, list):
                        v = demjson.encode(v, encoding='UTF-8')
                    row.append(v)
                    # print(f'{k} {t} {v}')
                sheet.append(row)
        else:
            row = ['']
            for i in range(1 if is_add_id else 0, len(ckey)):
                k = ckey[i]
                t = kid1[k]
                v = config.get(k)
                if t == 'dict' or t == 'list' or isinstance(v, dict) or isinstance(v, list):
                    v = json.dumps(v, ensure_ascii=False)  # demjson.encode(v, encoding='UTF-8')
                row.append(v)
            sheet.append(row)

    def _build_sheet2(sheet, config, title):
        sheet.title = title
        keys = []
        mkey = sorted(config['m'].keys())
        for k in mkey:
            keys.append(config['m'][k])
        print(f'{title}: {keys}')
        skey = []
        kid1 = []
        kid2 = []
        data = config['d']
        for i in range(len(mkey)-1, -1, -1):
            k = mkey[i]
            for item in data.values():
                if isinstance(item, dict):
                    val = item.get(k)
                    if val is None:
                        mkey.pop(i)
                        keys.pop(i)
                    else:
                        t = _guess_type(val)
                        kid1.append(t)
                        if t == 'dict' or t == 'list':
                            kid2.append('string')
                        else:
                            kid2.append(t)
                    break
                elif isinstance(item, list):
                    keys = ['id', 'item']
                    kid1 = ['int', 'list']
                    kid2 = ['int', 'string']
                    break
                else:
                    print('unkonw item type')
        kid1.reverse()
        kid2.reverse()
        # header
        sheet.append([title])
        row = ['CLIENT']
        row.extend(keys)
        sheet.append(row)
        row = ['']
        row.extend(kid2)
        sheet.append(row)
        row = ['SERVER']
        row.extend(skey)
        sheet.append(row)
        for rid in sorted(data.keys()):
            item = data[rid]
            row = ['']
            if isinstance(item, dict):
                for i in range(len(mkey)):
                    k = mkey[i]
                    if kid1[i] == 'dict' or kid1[i] == 'list':
                        row.append(demjson.encode(item.get(k), encoding='UTF-8'))
                    else:
                        row.append(item.get(k))
                sheet.append(row)
            elif isinstance(item, list):
                newitem = []
                for it in item:
                    newit = {}
                    for k in it:
                        newit[config['m'][k]] = it[k]
                    newitem.append(newit)
                row.append(rid)
                row.append(demjson.encode(newitem, encoding='UTF-8'))
                sheet.append(row)

    def _build_book1(book, config, title, roots=None):
        content, actived = None, False
        if roots:
            for r in roots:
                content = config.get(r)
                if content:
                    break
        content = content if content else config
        if isinstance(content, dict):
            for title in content.keys():
                print(f'--------------------- book:{title}')
                data = content[title]
                is_sub = True
                for v in data.values():
                    for sv in v.values():
                        if not isinstance(sv, dict) and not isinstance(sv, list):
                            is_sub = False
                            break
                if is_sub:
                    for k, v in data.items():
                        k = k.replace('?', '_')
                        if actived:
                            if book.__contains__(k):
                                b = book.get_sheet_by_name(k)
                            else:

                                b = book.create_sheet(k)
                        else:
                            actived = True
                            b = book.active
                        _build_sheet1(b, v, k)
                else:
                    if actived:
                        if book.__contains__(title):
                            b = book.get_sheet_by_name(title)
                        else:
                            b = book.create_sheet(title)
                    else:
                        actived = True
                        b = book.active
                    _build_sheet1(b, data, title)
        else:
            _build_sheet1(book.active, content, title)

    def _build_book2(book, config, title):
        if 'm' in config and 'd' in config:
            _build_sheet2(book.active, config, title)
            return
        isfirst = True
        for subkey in config:
            subobj = config[subkey]
            title = subkey.rstrip('_json')
            if 'm' in subobj and 'd' in subobj:
                if isfirst:
                    isfirst = False
                    _build_sheet2(book.active, subobj, title)
                else:
                    _build_sheet2(book.create_sheet(title), subobj, title)
            else:
                print('unexcept config format:', title)

    def _output1():
        src = '/Users/joli/Documents/AndroidStudio/DeviceExplorer/emulator-5554/data/data/com.lilithgames.hgame.cn/files/jsonc_dir'
        dst = '/Users/joli/Documents/AndroidStudio/DeviceExplorer/emulator-5554/data/data/com.lilithgames.hgame.cn/files/excel_dir'
        for fn in FS.walk_files(src, ewhites=['.json'], cut=len(src) + 1):
            if fn.endswith('languageDb.json'):
                continue
            file = os.path.join(src, fn)
            print('------------------------------------------', file)
            conf = demjson.decode_file(file)
            if not conf:
                print('empty config')
                continue
            name = FS.filename(file)
            wb = openpyxl.Workbook()
            # print(wb.get_sheet_names())
            _build_book1(wb, conf, name, ('ed', 'hg'))
            dstfile = os.path.join(dst, name + '.xlsx')
            FS.make_parent(dstfile)
            wb.save(dstfile)

    def _output2():
        json_ver = '/Users/joli/Downloads/fy/assets/game/resource/1_00.58_version.json'
        conf_src = '/Users/joli/Downloads/fy/assets/game/resource/config'
        conf_dst = '/Users/joli/Downloads/fy/assets/game/resource/config'
        ver_dict = {}
        for k, v in json.loads(FS.read_text(json_ver)).items():
            ver_dict[v] = k
        for f in FS.walk_files(conf_src, ewhites=['.json']):
            # f = '/Users/joli/Downloads/fy/assets/game/resource/config/4643a093.json'
            config = demjson.decode_file(f)
            if not config:
                print('empty config:' + f)
                continue
            vf = ver_dict[FS.filename(f)]
            df = os.path.join(conf_dst, vf[0: vf.rfind('.')] + '.xlsx')
            print(f)
            print(vf)
            print(df)
            wb = openpyxl.Workbook()
            if vf.startswith('global/') or 'error_code_data' in vf:
                print(config)
                _build_sheet1(wb.active, config, FS.filename(df), flat=True)
            else:
                _build_book1(wb, config, FS.filename(df))
            FS.make_parent(df)
            wb.save(df)
            # break

    _output1()
    # _output2()
    print('done')

def test_hack_apk():
    from simples.rob import RobAPK
    deapk = RobAPK.RobYQCR()
    # deapk.decompile()
    # deapk.reinstall()
    # deapk.extract_files()
    # deapk.pull_files()
    deapk.decrypt_res()
    # deapk.pull_files()

def main():
    pass
    # test_json2excel()
    # test_load_cdn_files()
    # test_hack_xor()
    # test_hack_xxtea()
    # test_hack_luac()
    # test_hack_zlib()
    # test_hack_respack()
    test_hack_apk()