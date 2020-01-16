# -*- coding: UTF-8 -*-

try :
    import sys
    import importlib
    importlib.reload(sys)
    sys.setdefaultencoding('utf8')
except:
    pass

import os
import re
import shutil

__author__ = 'Joli'

class Excel2:
    _ROW_DESC = 0
    _ROW_CKEY = 1  # CLIENT
    _ROW_CAST = 2
    _ROW_SKEY = 3  # SERVER
    _ROW_DATA = 4
    _COL_DATA = 1
    _COL_FLAG = 0  # END
    _FLAG_END = 'END'
    _illegals = ("\"", "\\", "/", "*", "'", "=", "-", "#", ";", "<", ">", "+", "%", "$", "(", ")", "%", "@", "!")

    class _SheetError:
        def __init__(self):
            self._err_vals = None
            self._err_keys = None
            self._err_sigs = None
            self._err_list = None

        def dump_errors(self, matrix):
            self._err_vals = []
            self._err_keys = []
            self._err_sigs = []
            self._err_list = [self._err_sigs, self._err_keys, self._err_vals]

            data_col_list = self._dump_data_columns(matrix)
            if len(data_col_list) > 0:
                self._dump_col_casts(matrix, data_col_list)
                self._dump_value_err(matrix, data_col_list)

            return self._err_list

        def _add_err_val(self, matrix, row, col, msg=''):
            self._err_vals.append('(%d,%d) - [%s] - %s' % (row, col, str(matrix[row][col]), msg))

        def _add_err_key(self, matrix, row, col, msg=''):
            self._err_keys.append('(%d,%d) - [%s] - %s' % (row, col, str(matrix[row][col]), msg))

        def _add_err_sig(self, matrix, row, col, msg=''):
            self._err_sigs.append('(%d,%d) - [%s] - %s' % (row, col, str(matrix[row][col]), msg))

        def _dump_data_columns(self, matrix):
            data_col_list = []
            for row in (Excel2._ROW_CKEY, Excel2._ROW_SKEY):
                row_len = len(matrix[row])
                for col in range(Excel2._COL_DATA, row_len):
                    key = Excel2._opt_col_key(matrix, row, col)
                    if key is None:
                        self._add_err_key(matrix, row, col, 'illegal key')
                        continue
                    if key == '':
                        continue
                    duplicate_col_list = []
                    for dup_col in range(col + 1, row_len, 1):
                        next_key = Excel2._opt_col_key(matrix, row, dup_col)
                        if key == next_key:
                            duplicate_col_list.append(dup_col)
                    if len(duplicate_col_list) > 0:
                        self._add_err_key(matrix, row, col, 'duplicate:%s' % str(duplicate_col_list))
                    if col not in data_col_list:
                        data_col_list.append(col)
            return data_col_list

        def _dump_col_casts(self, matrix, data_col_list):
            col_cast_dict = {}
            row = Excel2._ROW_CAST
            cast_row_cells = matrix[row]
            cast_row_len = len(cast_row_cells)
            for col in data_col_list:
                if col < cast_row_len:
                    cast = Excel2._opt_col_cast(matrix, col)
                    if cast is None:
                        self._add_err_sig(matrix, row, col, 'illegal cast')
                    else:
                        col_cast_dict[col] = cast
                else:
                    self._add_err_sig(matrix, row, col, 'empty cast')
            return col_cast_dict

        def _dump_value_err(self, matrix, data_col_list):
            for row in range(Excel2._ROW_DATA, len(matrix)):
                for col in data_col_list:
                    if Excel2._opt_cell_value(matrix, row, col) is not None:
                        continue
                    self._add_err_val(matrix, row, col, 'cast %s' % str(matrix[Excel2._ROW_CAST][col]))
                if matrix[row][Excel2._COL_FLAG].strip() == Excel2._FLAG_END:
                    break

    class SheetData:

        def __init__(self, sheet_name, matrix=None):
            self._sheet_name = sheet_name
            self._matrix = matrix

        def get_sheet_name(self):
            return self._sheet_name

        def get_matrix(self):
            return self._matrix

        def get_client_data(self):
            return self._get_data(Excel2._ROW_CKEY)

        def get_server_data(self):
            return self._get_data(Excel2._ROW_SKEY)

        def set_client_data(self, data):
            self._set_data(data, Excel2._ROW_CKEY)

        def set_server_data(self, data):
            self._set_data(data, Excel2._ROW_SKEY)

        def set_comment(self, row, col, msg):
            if row >= len(self._matrix):
                print('row out of range:', row)
                return
            row_data = self._matrix[row]
            if col >= len(row_data):
                print('col out of range:', col)
                return
            row_data[col] = msg

        def _get_data(self, key_row):
            data = {}
            mtx = self._matrix
            for i in range(Excel2._ROW_DATA, len(mtx)):
                flag = mtx[i][Excel2._COL_FLAG].strip()
                if i == Excel2._ROW_DATA and flag == Excel2._FLAG_END:
                    break #第一行有END，认为是空表。
                row_id = Excel2._opt_cell_value(mtx, i, Excel2._COL_DATA)
                if row_id is None:
                    continue
                config = {}
                for j in range(Excel2._COL_DATA + 1, len(mtx[i])):
                    key = Excel2._opt_col_key(mtx, key_row, j)
                    if not key:
                        continue
                    config[key] = Excel2._opt_cell_value(mtx, i, j)
                if len(config) > 0:
                    data[row_id] = config
                if flag == Excel2._FLAG_END:
                    break #遇到END，结束。
            return data

        def _set_data(self, data, key_row):
            key_list = self._matrix[key_row]
            row_list = list(data.keys())
            row_nums = len(row_list)
            for i in range(row_nums):
                row_id = row_list[i]
                row_data = self._find_row_data(Excel2._COL_DATA, row_id)
                if row_data is None:
                    row_data = []
                    self._safe_assign(row_data, Excel2._COL_FLAG, ('' if i+1 < row_nums else Excel2._FLAG_END))
                    self._safe_assign(row_data, Excel2._COL_DATA, row_id)
                    self._matrix.append(row_data) # add new line
                config = data[row_id]
                for key in config:
                    if key not in key_list:
                        print('miss key:', key)
                        continue
                    self._safe_assign(row_data, key_list.index(key), config[key])

        def _find_row_data(self, col, value):
            for i in range(Excel2._ROW_DATA, len(self._matrix)):
                row_data = self._matrix[i]
                if value == row_data[col]:
                    return row_data

        @staticmethod
        def _safe_assign(row_data, col, value):
            for j in range(col - len(row_data) + 1):
                row_data.append('')
            row_data[col] = value

        def init_matrix(self):
            self._matrix = []
            self._matrix.insert(Excel2._ROW_DESC, [self._sheet_name])
            self._matrix.insert(Excel2._ROW_CKEY, ['CLIENT'])
            self._matrix.insert(Excel2._ROW_CAST, [''])
            self._matrix.insert(Excel2._ROW_SKEY, ['SERVER'])

        def set_header(self, col, ckey='', cast='', skey='', desc=''):
            self._safe_assign(self._matrix[Excel2._ROW_DESC], col, desc)
            self._safe_assign(self._matrix[Excel2._ROW_CKEY], col, ckey)
            self._safe_assign(self._matrix[Excel2._ROW_CAST], col, cast)
            self._safe_assign(self._matrix[Excel2._ROW_SKEY], col, skey)


    @classmethod
    def _is_illegal_char(cls, chars):
        for c in cls._illegals:
            if c in chars:
                return True
        return False

    @classmethod
    def _opt_col_key(cls, matrix, row, col):
        key = matrix[row][col]
        if not isinstance(key, str):
            # print(col, 'non string key:', str(cast))
            return None
        key = key.strip()
        if key and cls._is_illegal_char(key):
            # print(col, 'illegal key')
            return None
        return key

    @classmethod
    def _opt_col_cast(cls, matrix, col):
        cast = matrix[cls._ROW_CAST][col]
        if not isinstance(cast, str):
            # print(col, 'non string cast:', str(cast))
            return None
        cast = cast.strip()
        if cast == '':
            # print(col, 'empty cast')
            return None
        if cls._is_illegal_char(cast):
            # print(col, 'illegal cast')
            return None
        return cast.lower()

    @classmethod
    def _opt_cell_value(cls, matrix, row, col):
        cast = cls._opt_col_cast(matrix, col)
        if 'int' == cast or 'float' == cast:
            float_val = cls._opt_float_value(matrix, row, col)
            if float_val is None:
                return None
            return float_val if 'float' == cast else int(float_val)
        cell = matrix[row][col]
        if cast == 'string':
            return str(cell)
        return cell

    @classmethod
    def _opt_float_value(cls, matrix, row, col):
        cell = matrix[row][col]
        try:
            return float(cell)
        except ValueError as e:
            # print(e)
            pass
        return None if cell.strip() else 0

    @classmethod
    def _is_configure(cls, sheet):
        if sheet.nrows <= cls._ROW_DATA or sheet.ncols <= cls._COL_DATA:
            return False
        v = sheet.cell_value(cls._ROW_CKEY, 0).strip()
        if 'CLIENT' != v:
            return False
        v = sheet.cell_value(cls._ROW_SKEY, 0).strip()
        if 'SERVER' != v:
            return False
        return True

    @classmethod
    def _create(cls, excel_path, sheet_data_list):
        import xlwt

        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_LEFT
        alignment.vert = xlwt.Alignment.VERT_TOP

        pattern = xlwt.Pattern()
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = 0

        font = xlwt.Font()
        font.colour_index = 5

        head_style = xlwt.XFStyle()
        head_style.font = font
        head_style.pattern = pattern
        head_style.alignment = alignment

        data_style = xlwt.XFStyle()
        data_style.alignment = alignment

        book = xlwt.Workbook(encoding='utf-8', style_compression=0)
        for sheet_data in sheet_data_list:
            sheet = book.add_sheet(sheet_data.get_sheet_name(), cell_overwrite_ok=False)
            matrix = sheet_data.get_matrix()
            for row in range(len(matrix)):
                row_data = matrix[row]
                style = head_style if row < cls._ROW_DATA else data_style
                for col in range(len(row_data)):
                    sheet.write(row, col, row_data[col], style)
        book.save(excel_path)

    @classmethod
    def _modify(cls, excel_path, sheet_data_list):
        from openpyxl import load_workbook
        old_excel = load_workbook(filename=excel_path)
        for sheet_data in sheet_data_list:
            sheet = old_excel[sheet_data.get_sheet_name()]
            matrix = sheet_data.get_matrix()
            for row in range(len(matrix)):
                row_data = matrix[row]
                for col in range(len(row_data)):
                    sheet.cell(row=row+1, column=col+1).value = row_data[col]
        path_info = os.path.split(excel_path)
        temp_path = os.path.join(path_info[0], 'temp_' + path_info[1])
        old_excel.save(temp_path)
        shutil.move(temp_path, excel_path)

    @classmethod
    def save(cls, excel_path, sheet_data_list):
        if not os.path.exists(excel_path):
            cls._create(excel_path, sheet_data_list)
        else:
            cls._modify(excel_path, sheet_data_list)

    @classmethod
    def read(cls, excel_path):
        import xlrd
        book = xlrd.open_workbook(excel_path)
        sheet_data_list = []
        for sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            if not cls._is_configure(sheet):
                continue
            matrix = []
            for row in range(sheet.nrows):
                row_data = []
                for col in range(sheet.ncols):
                    row_data.append(sheet.cell_value(row, col))
                matrix.append(row_data)
            sheet_data = cls.SheetData(sheet_name, matrix)
            sheet_data_list.append(sheet_data)
        return sheet_data_list

    @classmethod
    def print_excel_errors(cls, src_dir):
        sheet_error = cls._SheetError()
        key_error_log = open(os.path.join(src_dir, '..', 'key_error.txt'), 'w')
        val_error_log = open(os.path.join(src_dir, '..', 'val_error.txt'), 'w')
        sig_error_log = open(os.path.join(src_dir, '..', 'sig_error.txt'), 'w')
        err_log_files = [sig_error_log, key_error_log, val_error_log]
        for (root, dirs, files) in os.walk(src_dir):
            for file in files:
                if file[0] == '.':
                    continue
                sheet_path = os.path.join(root, file)
                print('sheet_path:', sheet_path)
                excel_records = []
                for sheet_data in cls.read(sheet_path):
                    print('---------------------------', sheet_data.get_sheet_name())
                    print('client:', sheet_data.get_client_data())
                    print('server:', sheet_data.get_server_data())
                    error_list = sheet_error.dump_errors(sheet_data.get_matrix())
                    for i in range(len(error_list)):
                        error_items = error_list[i]
                        if len(error_items) == 0:
                            continue
                        log_file = err_log_files[i]
                        if i not in excel_records:
                            excel_records.append(i)
                            log_file.writelines('\n\n' + sheet_path)
                        log_file.writelines('\n---------------------------')
                        log_file.writelines(sheet_data.get_sheet_name())
                        for err_log in error_items:
                            log_file.writelines('\n' + err_log)
        for log_file in err_log_files: log_file.close()
        print('done')

    @classmethod
    def create_excel_simple(cls, excel_path):
        simple_data = {1:{'k1':11, 'k2':12},2:{'k1':21, 'k2':22}}
        sheet_data = cls.SheetData('simple')
        sheet_data.init_matrix()
        sheet_data.set_header(1, 'id', 'int',   'id', 'ID')
        sheet_data.set_header(2, 'k1', 'string','',   'key1')
        sheet_data.set_header(3, 'k2', 'int',   '',   'key2')
        sheet_data.set_client_data(simple_data)
        cls.save(excel_path, [sheet_data])
        sheet_data = cls.read(excel_path)[0]
        print(sheet_data.get_client_data())

class Lang2:
    _LANG2_API = 'Lang2.getText(%d)'

    class Helper:
        _re_program_log = re.compile(
            r'((?<!-)--(\s*?)\[\[.*?\]\])|((--(?!\s*?\]\])).*?\n)|((print|dump)\s*?\(.*?\)[^,]*?\n)',
            re.I | re.S | re.M)
        _re_quote_block = re.compile(r'(\".*?(?<!\\)\")|(\'.*?(?<!\\)\')|(\[\[.*?\]\])', re.I | re.S | re.M)
        _re_remove_quto = re.compile(r'(^(\"|\'|\[\[))|((\"|\'|\]\])$)', re.I | re.S | re.M)
        _re_lang_target = re.compile(r'[\u4e00-\u9fa5]', re.I)

        # _re_cleanup_key = re.compile(r'\s+?', re.I|re.S |re.M)
        # @classmethod
        # def get_lang_key(cls, item):
        #     return cls._re_cleanup_key.sub('', item)

        @classmethod
        def hit_lang(cls, text):
            return cls._re_lang_target.search(text) is not None

        @classmethod
        def find_quote_blocks(cls, text):
            quot_list = []
            text = cls._re_program_log.sub('', text)
            info = cls._re_quote_block.findall(text)
            if info is None:
                return quot_list
            for items in info:
                for quot in items:
                    if quot != '':
                        quot_list.append(quot)
            # print(quot_list)
            return quot_list

        @classmethod
        def remove_quotes(cls, text):
            text = cls._re_remove_quto.sub('', text)
            # print('remove quote:', text)
            return text

    class LangData:
        def __init__(self, file_path, lang_list):
            self._file_path = file_path
            self._lang_list = lang_list
            self._file_text = None

        def set_file_text(self, file_text):
            self._file_text = file_text

        def get_file_text(self):
            return self._file_text

        def get_lang_list(self):
            return self._lang_list

        def get_file_path(self):
            return self._file_path

    def __init__(self):
        self._unicode_decode_errors = {}
        self._unicode_encode_errors = {}
        self._no_exists_lang_errors = {}
        self._new_lang_records_dict = {}

    def _read_string(self, path):
        with open(path, "rb") as f:
            b = f.read()
        if b is not None:
            try:
                return b.decode('utf-8')
            except UnicodeDecodeError as e:
                self._unicode_decode_errors[path] = e

    def _save_string(self, path, text):
        b = None
        try:
            b = text.encode('utf-8')
        except UnicodeEncodeError as e:
            self._unicode_encode_errors[path] = e
        if b is not None:
            with open(path, "wb") as f:
                f.write(b)

    @staticmethod
    def _is_ignore_path(root_dir, file_path, ignore_files):
        if ignore_files is not None:
            file_name = file_path[1 + len(root_dir):]
            for f in ignore_files:
                if file_name.startswith(f):
                    return True
        return False

    def _extract_file(self, file_path, lang_data_list, retain_text=False):
        file_text = self._read_string(file_path)
        if file_text is None:
            return

        lang_list = []
        quot_list = Lang2.Helper.find_quote_blocks(file_text)
        for quot in quot_list:
            if not Lang2.Helper.hit_lang(quot):
                continue
            if quot not in lang_list:
                lang_list.append(quot)

        if len(lang_list) > 0:
            lang_data = Lang2.LangData(file_path, lang_list)
            if retain_text:
                lang_data.set_file_text(file_text)
            lang_data_list.append(lang_data)

    def _do_extract(self, src_dir, ignore_files=None, retain_text=False):
        lang_data_list = []
        for (root, dirs, files) in os.walk(src_dir):
            for file in files:
                if ('.' == file[0]) or ('.lua' != file[len(file) - 4:]):
                    continue
                file_path = os.path.join(root, file)
                if self._is_ignore_path(src_dir, file_path, ignore_files):
                    continue
                self._extract_file(file_path, lang_data_list, retain_text)
        return lang_data_list

    def _do_replace(self, file_lang_list, lang_mapping, lang_api_fmt):
        for lang_data in file_lang_list:
            file_path = lang_data.get_file_path()
            file_text = lang_data.get_file_text()
            for lang_item in lang_data.get_lang_list():
                no_quot_lang = Lang2.Helper.remove_quotes(lang_item)
                langid = lang_mapping[no_quot_lang]
                if langid is None:
                    errors = self._no_exists_lang_errors.get(file_path)
                    if errors is None:
                        errors = []
                        self._no_exists_lang_errors[file_path] = errors
                    errors.append(lang_item)
                else:
                    file_text = file_text.replace(lang_item, (lang_api_fmt % int(langid)))
            self._save_string(file_path, file_text)

    def _read_lang_dict(self, lang_path):
        lang_dict = {}
        if not os.path.exists(lang_path):
            return lang_dict
        sheet_data_list = Excel2.read(lang_path)
        if len(sheet_data_list) == 0:
            return lang_dict
        sheet_data = sheet_data_list[0]
        return sheet_data.get_client_data()

    def _save_lang_config(self, lang_path, file_lang_list):
        lang_dict = self._read_lang_dict(lang_path)
        lang_mapping = {}
        update = self._update_lang_config(file_lang_list, lang_dict, lang_mapping)
        if update:
            sheet_data = Excel2.SheetData('ClientLang')
            sheet_data.init_matrix()
            sheet_data.set_header(1, 'id', 'int',    '', 'lang id')
            sheet_data.set_header(2, 'zh', 'string', '', '中文')
            sheet_data.set_header(3, 'en', 'string', '', '英文')
            sheet_data.set_client_data(lang_dict)
            Excel2.save(lang_path, [sheet_data])
        return lang_mapping

    def _update_lang_config(self, file_lang_list, lang_dict, lang_mapping):
        update = False
        for lang_id in lang_dict:
            lang_zone = lang_dict[lang_id]
            lang_mapping[lang_zone['zh']] = lang_id
        for lang_data in file_lang_list:
            for lang_item in lang_data.get_lang_list():
                lang_text = Lang2.Helper.remove_quotes(lang_item)
                if lang_text in lang_mapping:
                    continue
                lang_id = str(len(lang_mapping))
                lang_mapping[lang_text] = lang_id
                lang_dict[lang_id] = {'zh':lang_text}
                file_path = lang_data.get_file_path()
                new_items = self._new_lang_records_dict.get(file_path)
                if new_items is None:
                    new_items = []
                    self._new_lang_records_dict[file_path] = new_items
                new_items.append(lang_item)
                update = True
        return update

    def _print_phase(self, src_dir):
        exit_code = 0
        print_dir = os.path.join(src_dir, '..', os.path.basename(src_dir))

        new_lang_records_info_txt = print_dir + '_new_lang_records_info.txt'
        if len(self._new_lang_records_dict) > 0:
            f = open(new_lang_records_info_txt, 'w')
            for file in self._new_lang_records_dict:
                f.writelines(file + '\n')
                for lang_item in self._new_lang_records_dict[file]:
                    f.writelines(lang_item + '\n')
                f.writelines('\n\n')
            f.close()
        else:
            if os.path.exists(new_lang_records_info_txt):
                os.remove(new_lang_records_info_txt)

        unicode_decode_errors_txt = print_dir + '_unicode_decode_errors.txt'
        if len(self._unicode_decode_errors) > 0:
            f = open(unicode_decode_errors_txt, 'w')
            for file in self._unicode_decode_errors:
                f.writelines(file + '\n')
            f.close()
            exit_code = 1
        else:
            if os.path.exists(unicode_decode_errors_txt):
                os.remove(unicode_decode_errors_txt)

        unicode_encode_errors_txt = print_dir + '_unicode_encode_errors.txt'
        if len(self._unicode_encode_errors) > 0:
            f = open(unicode_encode_errors_txt, 'w')
            for file in self._unicode_encode_errors:
                f.writelines(file + '\n')
            f.close()
            exit_code = 2
        else:
            if os.path.exists(unicode_encode_errors_txt):
                os.remove(unicode_encode_errors_txt)

        no_exists_lang_errors_txt = print_dir + '_no_exists_lang_errors.txt'
        if len(self._no_exists_lang_errors) > 0:
            f = open(no_exists_lang_errors_txt, 'w')
            for file in self._no_exists_lang_errors:
                f.writelines(file + '\n')
                for lang_item in self._no_exists_lang_errors[file]:
                    f.writelines(lang_item + '\n')
                f.writelines('\n\n')
            f.close()
            exit_code = 3
        else:
            if os.path.exists(no_exists_lang_errors_txt):
                os.remove(no_exists_lang_errors_txt)
        return exit_code

    def execute(self, lang_path, src_dir, ignore_files=None, lang_api_fmt=_LANG2_API):
        print('Lang2 extracting ...')
        use_replace = (lang_api_fmt is not None) or (0 == len(lang_api_fmt))
        file_lang_list = self._do_extract(src_dir, ignore_files, use_replace)

        if len(file_lang_list) > 0:
            lang_mapping = self._save_lang_config(lang_path, file_lang_list)
            if use_replace:
                print('Lang2 replacing ...')
                self._do_replace(file_lang_list, lang_mapping, lang_api_fmt)

        exit_code = self._print_phase(src_dir)
        print('Lang2 done')
        return exit_code

class Client2:

    def __init__(self, sys_args):
        from optparse import OptionParser
        parser = OptionParser(description='client tool')
        parser.add_option('--sheet2lua',
                          dest='sheet2lua',
                          action='store_true',
                          help='export excel sheet to lua')
        parser.add_option('--validate-sheet',
                          dest='validate_sheet',
                          action='store_true',
                          help='list excel sheet errors')
        parser.add_option('--list-language',
                          dest='list_language',
                          action='store_true',
                          help='list program language')
        parser.add_option('--translate-language',
                          dest='translate_language',
                          action='store_true',
                          help='translate program language')
        parser.add_option('-i', '--src',
                          dest='src_path',
                          type='string',
                          default=None,
                          help='src path for task')
        parser.add_option('-o', '--dst',
                          dest='dst_path',
                          type='string',
                          default=None,
                          help='dst path for task')
        (options, args) = parser.parse_args(sys_args[1:])
        if not options:
            pass
        elif options.list_language:
            self.__task_list_language(options.src_path, options.dst_path)
        elif options.translate_language:
            self.__task_translate_language(options.src_path, options.dst_path)
        elif options.validate_sheet:
            self.__task_validate_sheet(options.src_path)
        elif options.sheet2lua:
            self.__task_sheet2lua(options.src_path, options.dst_path)

    def __task_list_language(self, source_dir, lang_excel):
        if not source_dir:
            source_dir = '/Users/joli/Desktop/test/lua/dazhanguo_fanyi/src'
        if not lang_excel:
            lang_excel = '/Users/joli/Desktop/test/lua/dazhanguo_fanyi/ClientLang.xlsx'
        exit_code = Lang2().execute(lang_excel, source_dir, ['config/', 'lang/'], '')
        sys.exit(exit_code)

    def __task_translate_language(self, source_dir, lang_excel):
        if not source_dir:
            source_dir = '/Users/joli/Desktop/test/lua/dazhanguo_fanyi/src'
        if not lang_excel:
            lang_excel = '/Users/joli/Desktop/test/lua/dazhanguo_fanyi/ClientLang.xlsx'
        exit_code = Lang2().execute(lang_excel, source_dir, ['config/', 'lang/'])
        sys.exit(exit_code)

    def __task_validate_sheet(self, excel_dir):
        if not excel_dir:
            excel_dir = '/Users/joli/Desktop/test/excel/config'
        Excel2.print_excel_errors(excel_dir)

    def __task_sheet2lua(self, excel_dir, lua_dir):
        if not excel_dir:
            excel_dir = '/Users/joli/Desktop/test/excel/config'
        if not lua_dir:
            lua_dir = ''
        # Excel2.create_excel_simple('/Users/joli/Desktop/test/excel/test/simple.xlsx')

if __name__ == '__main__':
    Client2(sys.argv)